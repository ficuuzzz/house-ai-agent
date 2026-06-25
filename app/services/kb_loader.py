import ast
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[2]
KB_FILE_PATH = BASE_DIR / "data" / "KnowledgeBase.xlsx"

BOOLEAN_CONDITIONS = {
    "has_gas",
    "has_generator",
    "has_pool",
    "has_basement",
    "has_plot",
    "has_fireplace",
}


def normalize_header(value: Any) -> str:
    return str(value).strip().lower()


def normalize_condition(condition: str) -> str:
    condition = condition.strip()

    if condition in BOOLEAN_CONDITIONS:
        return f"{condition}=true"

    return condition


def parse_conditions(value: Any) -> list[str]:
    if value is None:
        return []

    if isinstance(value, list):
        return [normalize_condition(str(item)) for item in value if str(item).strip()]

    text = str(value).strip()

    if text == "" or text == "[]":
        return []

    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, list):
                return [
                    normalize_condition(str(item))
                    for item in parsed
                    if str(item).strip()
                ]
        except Exception:
            pass

    if "," in text:
        return [
            normalize_condition(item)
            for item in text.split(",")
            if item.strip()
        ]

    if ";" in text:
        return [
            normalize_condition(item)
            for item in text.split(";")
            if item.strip()
        ]

    return [normalize_condition(text)]


def load_knowledge_base_from_excel() -> list[dict]:
    if not KB_FILE_PATH.exists():
        raise FileNotFoundError(
            f"KnowledgeBase file not found: {KB_FILE_PATH}"
        )

    workbook = load_workbook(KB_FILE_PATH)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]

    items = []

    for row in rows[1:]:
        raw_item = dict(zip(headers, row))

        if not raw_item.get("kb_id"):
            continue

        item = {
            "kb_id": str(raw_item.get("kb_id")),
            "month": raw_item.get("month"),
            "season": raw_item.get("season"),
            "category": raw_item.get("category"),
            "subcategory": raw_item.get("subcategory"),
            "title": raw_item.get("title"),
            "task_description": raw_item.get("task_description"),
            "instructions": raw_item.get("instructions"),
            "purpose": raw_item.get("purpose"),
            "conditions": parse_conditions(raw_item.get("conditions")),
            "can_do_self": raw_item.get("can_do_self"),
            "priority": raw_item.get("priority"),
            "source_url": raw_item.get("source_url"),
        }

        items.append(item)

    return items